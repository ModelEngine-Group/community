#!/usr/bin/env python3
"""
ModelEngine GitHub 贡献者代码量统计脚本
- 抓取 GitHub (ModelEngine-Group) 所有公开仓库的贡献者数据
- 通过 /contributors 分页获取全部贡献者（解决 /stats/contributors 最多100条的限制）
- 输出 Excel 到指定目录
"""

import json
import os
import sys
import time
import http.client
from datetime import datetime, timedelta, timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# Read GitHub token from file
TOKEN_FILE = "/root/.github_token"
USERID = ""
def get_token():
    """Read GitHub token from file."""
    if not os.path.exists(TOKEN_FILE):
        raise FileNotFoundError(f"Token file not found: {TOKEN_FILE}")
    with open(TOKEN_FILE, "r") as f:
        return f.read().strip()

GITHUB_ORG = "ModelEngine-Group"
OUTPUT_DIR = "/root/{USERID}/star"
os.makedirs(OUTPUT_DIR, exist_ok=True)


def get_headers():
    token = os.environ.get("GITHUB_TOKEN", "")
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/vnd.github.v3+json"}
    if token:
        headers["Authorization"] = f"token {token}"
    return headers


def github_request(method, url, max_retries=3):
    """Make a request to GitHub API and return (status, data, link_header)."""
    headers = get_headers()
    for attempt in range(max_retries):
        try:
            conn = http.client.HTTPSConnection("api.github.com", timeout=30)
            conn.request(method, url, headers=headers)
            resp = conn.getresponse()
            raw = resp.read().decode()
            status = resp.status
            link = resp.getheader("Link", "")
            conn.close()
            if status == 403:
                wait = min(60 * (2 ** attempt), 180)
                print(f"    🔒 Rate limited, waiting {wait}s...")
                time.sleep(wait)
                continue
            if raw:
                data = json.loads(raw)
            else:
                data = []
            return status, data, link
        except Exception as e:
            wait = 5 * (attempt + 1)
            time.sleep(wait)
    return 0, [], ""


def fetch_github_repos():
    """Fetch all public repos from GitHub org, excluding specified repos."""
    all_repos = []
    page = 1
    while True:
        url = f"/orgs/{GITHUB_ORG}/repos?per_page=100&sort=updated&page={page}"
        status, data, _ = github_request("GET", url)
        if status != 200 or not data:
            break
        # Exclude Label-Studio (fork repo) and deer-flow
        data = [r for r in data if r["name"] not in ("Label-Studio", "deer-flow")]
        all_repos.extend(data)
        if len(data) < 100:
            break
        page += 1
    return all_repos


def fetch_all_contributors(repo_name):
    """Fetch ALL contributors for a repo using /contributors endpoint with pagination."""
    all_contributors = []
    page = 1
    while True:
        url = f"/repos/{GITHUB_ORG}/{repo_name}/contributors?per_page=100&page={page}&anon=1"
        status, data, link = github_request("GET", url)
        if status != 200 or not data:
            break
        all_contributors.extend(data)
        if len(data) < 100:
            break
        page += 1
        time.sleep(0.3)
    return all_contributors


def fetch_stats_batch(repo_name, max_retries=5):
    """Fetch stats via /stats/contributors (max 100 contributors with code stats)."""
    url = f"/repos/{GITHUB_ORG}/{repo_name}/stats/contributors"
    for attempt in range(max_retries):
        status, data, _ = github_request("GET", url)
        if status == 200:
            result = {}
            for c in data:
                author = c.get("author")
                if author is None:
                    continue
                login = author.get("login", f"anon_{author.get('id', 'unknown')}")
                week_stats = c.get("weeks", [])
                additions = sum(w.get("a", 0) for w in week_stats)
                deletions = sum(w.get("d", 0) for w in week_stats)
                total = c.get("total", 0)
                result[login] = {"additions": additions, "deletions": deletions, "total": total}
            return result

        if status == 202:
            wait = min(3 * (2 ** attempt), 30)
            print(f"    ⏳ Stats computing (202), waiting {wait}s...")
            time.sleep(wait)
            continue

        if status in (204, 404):
            return {}

        break
    return None


def get_all_contributor_stats(repo_name):
    """Get stats for ALL contributors of a repo.
    
    Strategy:
    1. Paginate /contributors to get ALL contributor names (no limit)
    2. Fetch /stats/contributors batch (has code stats, max 100)
    3. Match contributors: use batch stats if available, otherwise 0
    """
    # Step 1: Get all contributors via pagination
    all_contributors = fetch_all_contributors(repo_name)
    print(f"  Found {len(all_contributors)} total contributors (via pagination)")

    # Step 2: Get batch stats (may have up to 100)
    batch_stats = fetch_stats_batch(repo_name)
    if batch_stats is None:
        batch_stats = {}

    # Step 3: Merge
    result = []
    for c in all_contributors:
        login = c.get("login")
        if login is None:
            # Anonymous contributor
            name = c.get("name", "anonymous")
            if name and name != "anonymous":
                login = name
            else:
                login = f"anon_{c.get('id', 'unknown')}"
        
        stats = batch_stats.get(login, {})
        result.append({
            "login": login,
            "additions": stats.get("additions", 0),
            "deletions": stats.get("deletions", 0),
            "total": stats.get("total", 0),
        })

    return result


def save_to_excel(rows, output_path, update_time):
    """Save contributor data to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "贡献者代码量统计"

    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    name_font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    title_font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
    time_font = Font(name="Microsoft YaHei", bold=True, size=11, color="C00000")
    ws.cell(row=1, column=1, value="ModelEngine GitHub 贡献者代码量统计").font = title_font
    ws.cell(row=2, column=1, value=f"数据更新时间：{update_time}").font = time_font

    headers = ["#", "仓库 Repo", "贡献者 ID", "增加代码量 (+)", "删除代码量 (-)", "提交次数"]
    col_widths = [5, 35, 25, 18, 18, 12]
    header_row = 3
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    rank = 0
    prev_repo = None
    repo_toggle = False
    for i, row in enumerate(rows, header_row + 1):
        if row["repo"] != prev_repo:
            repo_toggle = not repo_toggle
        prev_repo = row["repo"]

        rank += 1
        values = [
            rank, row["repo"], row["contributor"],
            row["additions"], row["deletions"], row["total_commits"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = name_font if col in (2, 3) else data_font
            cell.alignment = center if col not in (2, 3) else left
            cell.border = thin_border
            if col == 4 and val > 0:
                cell.font = Font(name="Calibri", size=10, color="006100")
            elif col == 5 and val > 0:
                cell.font = Font(name="Calibri", size=10, color="9C0006")
            if repo_toggle:
                cell.fill = alt_fill

    ws.freeze_panes = "A4"
    wb.save(output_path)
    return output_path


def main():
    now_utc = datetime.now(timezone.utc)
    now_bj_dt = now_utc + timedelta(hours=8)
    update_time = now_bj_dt.strftime("%Y-%m-%d %H:%M:%S (北京时间)")
    timestamp = now_utc.strftime("%Y%m%d_%H%M%S")

    print(f"[{update_time}] Starting contributor stats collection...")

    print("Fetching GitHub repos...")
    repos = fetch_github_repos()
    print(f"  Found {len(repos)} repos")

    all_rows = []
    for idx, repo in enumerate(repos, 1):
        repo_name = repo["name"]
        print(f"\n[{idx}/{len(repos)}] Fetching stats for: {repo_name}...")
        contributors = get_all_contributor_stats(repo_name)
        print(f"  Got {len(contributors)} contributor records")

        for c in contributors:
            all_rows.append({
                "repo": repo_name,
                "contributor": c["login"],
                "additions": c["additions"],
                "deletions": c["deletions"],
                "total_commits": c["total"],
            })

        # Pause between repos
        if idx < len(repos):
            time.sleep(2)

    all_rows.sort(key=lambda r: (r["repo"], -r["additions"]))

    filename = f"modelengine_contributor_stats_{timestamp}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    save_to_excel(all_rows, output_path, update_time)
    print(f"\nSaved to: {output_path}")

    print(f"\nSummary:")
    print(f"  Total repos: {len(repos)}")
    print(f"  Total contributor entries: {len(all_rows)}")
    total_add = sum(r["additions"] for r in all_rows)
    total_del = sum(r["deletions"] for r in all_rows)
    print(f"  Total additions: {total_add:,}")
    print(f"  Total deletions: {total_del:,}")

    contributor_totals = {}
    for r in all_rows:
        name = r["contributor"]
        if name not in contributor_totals:
            contributor_totals[name] = {"add": 0, "del": 0, "commits": 0}
        contributor_totals[name]["add"] += r["additions"]
        contributor_totals[name]["del"] += r["deletions"]
        contributor_totals[name]["commits"] += r["total_commits"]

    top = sorted(contributor_totals.items(), key=lambda x: -x[1]["add"])[:10]
    print(f"\nTop 10 contributors (by additions across all repos):")
    for name, stats in top:
        print(f"  {name}: +{stats['add']:,} / -{stats['del']:,} ({stats['commits']} commits)")

    print(f"\n[{update_time}] Done.")


if __name__ == "__main__":
    main()
