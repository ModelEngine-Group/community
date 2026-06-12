#!/usr/bin/env python3
"""
ModelEngine 跨平台 Star 统计脚本
- 抓取 GitHub (ModelEngine-Group) 和 GitCode (ModelEngine) 所有公开仓库的 star 数
- 按仓库名称合并，同名仓库 star 相加
- 输出 Excel 到 workspace 目录
"""

import json
import os
import urllib.request
import sys
from datetime import datetime, timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

GITHUB_ORG = "ModelEngine-Group"
GITCODE_ORG = "ModelEngine"
OUTPUT_DIR = "/root/zgqhope/star"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Known forks on GitCode that should be excluded from totals
GITCODE_FORKS = {"dingo", "data-juicer"}

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"


def fetch_github_repos():
    """Fetch all public repos from GitHub org."""
    base = f"https://api.github.com/orgs/{GITHUB_ORG}/repos?per_page=100&sort=updated"
    all_repos = []
    page = 1
    while True:
        url = f"{base}&page={page}"
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode())
        if not data:
            break
        all_repos.extend(data)
        if len(data) < 100:
            break
        page += 1

    return {r["name"]: r["stargazers_count"] for r in all_repos}


def fetch_gitcode_repos():
    """Fetch all public repos from GitCode/AtomGit API."""
    url = f"https://atomgit.com/api/v5/orgs/{GITCODE_ORG}/repos?per_page=100"
    all_repos = []
    page = 1
    while True:
        page_url = f"{url}&page={page}"
        req = urllib.request.Request(page_url, headers={"User-Agent": UA})
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode())
        if not data:
            break
        all_repos.extend(data)
        if len(data) < 100:
            break
        page += 1

    return {
        r["name"]: {
            "stars": r["stargazers_count"],
            "is_fork": r.get("fork", False) or r["name"] in GITCODE_FORKS,
            "status": r.get("status", ""),
        }
        for r in all_repos
    }


def merge_and_sort(github, gitcode):
    """Merge repos by name, sort by total stars descending."""
    all_names = set(github.keys()) | set(gitcode.keys())

    rows = []
    for name in all_names:
        gh_stars = github.get(name, 0)
        gc_info = gitcode.get(name, {"stars": 0, "is_fork": False, "status": ""})
        gc_stars = gc_info["stars"] if isinstance(gc_info, dict) else 0
        is_fork = gc_info["is_fork"] if isinstance(gc_info, dict) else False
        gc_status = gc_info.get("status", "") if isinstance(gc_info, dict) else ""
        total = gh_stars + gc_stars

        # Determine source
        sources = []
        if gh_stars > 0:
            sources.append(f"GitHub {gh_stars}")
        if gc_stars > 0:
            sources.append(f"GitCode {gc_stars}")

        rows.append({
            "name": name,
            "github_stars": gh_stars,
            "gitcode_stars": gc_stars,
            "total": total,
            "source": " | ".join(sources) if sources else "—",
            "is_fork": is_fork,
            "gc_status": gc_status,
        })

    # Sort by total stars descending
    rows.sort(key=lambda r: -r["total"])
    return rows


def save_to_excel(rows, output_path, update_time):
    """Save merged data to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Star 统计"

    # Styles
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    name_font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row with timestamp
    title_font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
    time_font = Font(name="Microsoft YaHei", bold=True, size=11, color="C00000")
    ws.cell(row=1, column=1, value="ModelEngine 跨平台 Star 统计").font = title_font
    ws.cell(row=2, column=1, value=f"数据更新时间：{update_time}").font = time_font

    # Headers
    headers = ["#", "仓库名称", "GitHub ⭐", "GitCode ⭐", "合计 ⭐", "来源", "备注"]
    col_widths = [5, 35, 12, 12, 12, 35, 20]
    header_row = 3
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    rank = 0
    for i, row in enumerate(rows, header_row + 1):
        if row["is_fork"]:
            continue  # Skip forks from main count
        rank += 1
        alt = i % 2 == 0

        values = [
            rank,
            row["name"],
            row["github_stars"],
            row["gitcode_stars"],
            row["total"],
            row["source"],
            "fork" if row["is_fork"] else "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = name_font if col == 2 else data_font
            cell.alignment = center if col != 2 else left
            cell.border = thin_border
            if alt:
                cell.fill = alt_fill

    # Summary section
    summary_row = header_row + len(rows) + 3
    non_fork_rows = [r for r in rows if not r["is_fork"]]
    total_stars = sum(r["total"] for r in non_fork_rows)
    dual_count = sum(1 for r in rows if not r["is_fork"] and r["github_stars"] > 0 and r["gitcode_stars"] > 0)
    gh_only = sum(1 for r in rows if not r["is_fork"] and r["github_stars"] > 0 and r["gitcode_stars"] == 0)
    gc_only = sum(1 for r in rows if not r["is_fork"] and r["github_stars"] == 0 and r["gitcode_stars"] > 0)

    ws.cell(row=summary_row, column=1, value="汇总统计").font = Font(bold=True, size=12, color="2F5496")
    ws.cell(row=summary_row + 1, column=1, value=f"总仓库数（去重）: {len(non_fork_rows)} 个").font = data_font
    ws.cell(row=summary_row + 2, column=1, value=f"双平台共有: {dual_count} 个").font = data_font
    ws.cell(row=summary_row + 3, column=1, value=f"仅 GitHub: {gh_only} 个").font = data_font
    ws.cell(row=summary_row + 4, column=1, value=f"仅 GitCode: {gc_only} 个").font = data_font
    ws.cell(row=summary_row + 5, column=1, value=f"合计总 Star: {total_stars:,} ⭐").font = Font(bold=True, size=11, color="C00000")

    # Freeze panes
    ws.freeze_panes = "A4"

    wb.save(output_path)
    return output_path


def main():
    # Timestamp in Beijing time
    now_utc = datetime.now(timezone.utc)
    now_bj = now_utc.strftime("%Y-%m-%d %H:%M:%S")  # UTC for display
    # Calculate Beijing time (UTC+8)
    from datetime import timedelta
    now_bj_dt = now_utc + timedelta(hours=8)
    update_time = now_bj_dt.strftime("%Y-%m-%d %H:%M:%S (北京时间)")
    timestamp = now_utc.strftime("%Y%m%d_%H%M%S")

    print(f"[{update_time}] Starting star stats collection...")

    # Fetch data
    print("Fetching GitHub repos...")
    github = fetch_github_repos()
    print(f"  Found {len(github)} repos")

    print("Fetching GitCode repos...")
    gitcode = fetch_gitcode_repos()
    print(f"  Found {len(gitcode)} repos")

    # Merge and sort
    rows = merge_and_sort(github, gitcode)

    # Save to Excel
    filename = f"modelengine_star_stats_{timestamp}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    save_to_excel(rows, output_path, update_time)
    print(f"\nSaved to: {output_path}")

    # Print summary
    non_fork = [r for r in rows if not r["is_fork"]]
    total_stars = sum(r["total"] for r in non_fork)
    print(f"\nSummary:")
    print(f"  Repos (dedup, excl forks): {len(non_fork)}")
    print(f"  Total stars: {total_stars:,}")

    # Top 5
    print(f"\nTop 5 by total stars:")
    for r in non_fork[:5]:
        print(f"  {r['name']}: {r['total']:,} ⭐ (GitHub {r['github_stars']:,} + GitCode {r['gitcode_stars']:,})")

    print(f"\n[{update_time}] Done.")


if __name__ == "__main__":
    main()
